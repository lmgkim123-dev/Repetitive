from __future__ import annotations

import re
import tempfile
from datetime import datetime
from io import BytesIO
from math import ceil
from pathlib import Path
from typing import Iterable

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from src.exporter import export_dataframes, export_excel
except ImportError:
    from src.exporter import export_excel
    export_dataframes = None
from src.pipeline import run_pipeline_v6

try:
    from src.piping_repeat_builder import (
        build_piping_repeat_report_dataframe,
        build_piping_summary_sheet,
        load_piping_replacement_occurrences,
    )
except ModuleNotFoundError:
    build_piping_repeat_report_dataframe = None
    build_piping_summary_sheet = None
    load_piping_replacement_occurrences = None
from src.task_builder import (
    CATEGORY_ORDER,
    build_category_extract_dataframe,
    build_equipment_summary_dataframe,
    categorize_event,
)

st.set_page_config(page_title="반복정비 과제 후보 생성 Tool", layout="wide")
st.title("반복정비 과제 후보 생성 Tool")
st.caption("고정장치와 배관을 분리 분석하여 반복 정비 후보를 도출합니다.")


# -------------------------------------------------
# common helpers
# -------------------------------------------------
def save_uploaded_files(files: Iterable) -> Path:
    temp_dir = Path(tempfile.mkdtemp(prefix="repeat_task_v6_"))
    for up in files:
        out_path = temp_dir / Path(up.name).name
        out_path.write_bytes(up.getbuffer())
    return temp_dir


def _truncate_multiline(text: str, max_len: int = 140) -> str:
    text = str(text or "").replace("\n", " / ")
    return text[:max_len] + ("..." if len(text) > max_len else "")


_GENERATED_FIXED_RESULT_RE = re.compile(r"반복정비_고정장치_(?:필터결과|조치요약|과제후보).*\.xls", re.I)
_GENERATED_FIXED_SHEETS = {"과제후보_등록형식", "카테고리별_발췌", "연도별_정비이벤트"}


def _looks_like_generated_fixed_result_file(uploaded_file) -> bool:
    name = Path(uploaded_file.name).name
    if _GENERATED_FIXED_RESULT_RE.search(name):
        return True
    if Path(name).suffix.lower() not in {".xlsx", ".xls", ".xlsm", ".xlsb"}:
        return False
    try:
        excel = pd.ExcelFile(BytesIO(uploaded_file.getbuffer()))
        if _GENERATED_FIXED_SHEETS.issubset(set(excel.sheet_names)):
            return True
    except Exception:
        pass
    return False


def split_fixed_input_files(files: list) -> tuple[list, list[str]]:
    usable = []
    skipped = []
    for uploaded_file in files:
        if _looks_like_generated_fixed_result_file(uploaded_file):
            skipped.append(uploaded_file.name)
        else:
            usable.append(uploaded_file)
    return usable, skipped


_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_CENTER_HEADERS = {
    "NO", "Equipment No", "발생구분", "발생년도수", "발생년도", "발췌 Category", "검토필요여부",
    "year", "event_year", "event_date", "source_count", "confidence", "발생횟수", "최근발생일", "source_type",
}
_WIDE_TEXT_HEADERS = {
    "설비명", "TA 조치사항", "추후 권고사항", "제목", "상세 내용", "source_files",
    "finding_location", "finding_damage", "finding_measurement", "action_type", "action_detail", "recommendation",
    "evidence_summary", "repeat_reason", "action_cluster", "location_cluster", "damage_cluster", "sources",
    "titles", "details", "출처", "대표조치", "상세이력", "검토메모", "title", "detail", "exclude_reason",
}
_FIXED_WIDTHS = {
    "NO": 7, "Equipment No": 15, "설비명": 30, "발생구분": 13, "발생년도수": 11, "발생년도": 18,
    "발췌 Category": 18, "TA 조치사항": 60, "추후 권고사항": 60, "제목": 42,
    "상세 내용": 92, "검토필요여부": 12, "equipment_no": 15, "equipment_name": 30, "year": 10,
    "source_files": 32, "finding_location": 24, "finding_damage": 24, "finding_measurement": 22,
    "action_type": 24, "action_detail": 60, "recommendation": 60, "evidence_summary": 80,
    "action_cluster": 20, "location_cluster": 20, "damage_cluster": 20, "repeat_reason": 48, "confidence": 10,
    "Line Number": 18, "발생횟수": 10, "최근발생일": 14, "출처": 24, "대표조치": 44, "상세이력": 66,
    "검토메모": 38, "line_no": 18, "event_date": 14, "event_year": 10, "sources": 24, "source_count": 10,
    "titles": 40, "details": 72, "source_type": 16, "title": 30, "detail": 78, "exclude_reason": 30,
}


def _fmt_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return str(value)


def _estimate_wrapped_lines(text: str, width: int) -> int:
    if not text:
        return 1
    usable = max(14, width - 2)
    count = 0
    for raw in text.splitlines() or [text]:
        line = raw.strip() or " "
        count += max(1, ceil(len(line) / usable))
    return count


def _estimate_col_width(ws, col_idx: int, header: str) -> float:
    if header in _FIXED_WIDTHS:
        return _FIXED_WIDTHS[header]
    max_len = len(header)
    for row_idx in range(2, min(ws.max_row, 150) + 1):
        txt = _fmt_text(ws.cell(row=row_idx, column=col_idx).value)
        max_len = max(max_len, max((len(x) for x in txt.splitlines()), default=0))
    if header in _WIDE_TEXT_HEADERS:
        return max(24, min(64, int(max_len * 1.05) + 2))
    return max(10, min(26, int(max_len * 1.1) + 2))


def _apply_readable_excel_format(output_path: Path) -> None:
    wb = load_workbook(output_path)
    for ws in wb.worksheets:
        if ws.max_row == 0 or ws.max_column == 0:
            continue
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.zoomScale = 90
        headers = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            header = _fmt_text(cell.value)
            headers.append(header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = _estimate_col_width(ws, col_idx, header)
        ws.row_dimensions[1].height = 26

        for row_idx in range(2, ws.max_row + 1):
            line_est = 1
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                text = _fmt_text(cell.value)
                width = int(ws.column_dimensions[get_column_letter(col_idx)].width or 12)
                cell.alignment = Alignment(
                    horizontal="center" if header in _CENTER_HEADERS else "left",
                    vertical="top",
                    wrap_text=True,
                )
                if header in _WIDE_TEXT_HEADERS or len(text) > 24 or "\n" in text:
                    line_est = max(line_est, _estimate_wrapped_lines(text, width))
            ws.row_dimensions[row_idx].height = min(220, max(22, 17 * line_est + 4))
    wb.save(output_path)


# -------------------------------------------------
# fixed equipment helpers
# -------------------------------------------------
def build_task_display_df(task_df: pd.DataFrame) -> pd.DataFrame:
    if task_df is None or task_df.empty:
        return pd.DataFrame()
    display_df = task_df.copy()
    display_df["TA 조치 요약"] = display_df.get("TA 조치사항", "").fillna("").astype(str).apply(lambda x: _truncate_multiline(x, 160))
    display_df["권고 요약"] = display_df.get("추후 권고사항", "").fillna("").astype(str).apply(lambda x: _truncate_multiline(x, 120))
    preferred_cols = [
        "NO", "Equipment No", "설비명", "발생구분", "발생년도수", "발생년도",
        "발췌 Category", "TA 조치 요약", "권고 요약",
    ]
    return display_df[[c for c in preferred_cols if c in display_df.columns]].copy()


def build_events_df(all_events: list) -> pd.DataFrame:
    rows = []
    for e in all_events:
        rows.append({
            "equipment_no": e.equipment_no,
            "equipment_name": e.equipment_name,
            "year": e.report_year,
            "정비Category": ", ".join(categorize_event(e)),
            "finding_location": e.finding_location,
            "finding_damage": e.finding_damage,
            "action_type": e.action_type,
            "action_detail": e.action_detail,
            "recommendation": e.recommendation,
            "evidence_summary": e.evidence_summary,
        })
    return pd.DataFrame(rows)


def build_cases_df(repeat_cases: list) -> pd.DataFrame:
    rows = []
    for c in repeat_cases:
        rows.append({
            "equipment_no": c.equipment_no,
            "설비명": c.equipment_name,
            "발생년도수": len(sorted(set(c.years))),
            "발생년도": ", ".join(map(str, sorted(set(c.years)))),
            "action_cluster": c.action_cluster,
            "location_cluster": c.location_cluster,
            "damage_cluster": c.damage_cluster,
            "repeat_reason": c.repeat_reason,
            "confidence": f"{c.confidence:.2f}",
        })
    return pd.DataFrame(rows)


def filter_task_df(task_df: pd.DataFrame, selected_categories: list[str], min_year_count: int) -> pd.DataFrame:
    if task_df is None or task_df.empty:
        return pd.DataFrame(columns=task_df.columns if task_df is not None else [])
    filtered = task_df.copy()
    filtered = filtered[pd.to_numeric(filtered["발생년도수"], errors="coerce").fillna(0) >= min_year_count].copy()
    if selected_categories:
        filtered = filtered[filtered["발췌 Category"].astype(str).isin(selected_categories)].copy()
    filtered = filtered.reset_index(drop=True)
    filtered["NO"] = range(1, len(filtered) + 1)
    return filtered


def filter_related_objects(filtered_task_df: pd.DataFrame, repeat_cases: list, all_events: list):
    if filtered_task_df is None or filtered_task_df.empty:
        return [], []
    keys = {(str(r["Equipment No"]), str(r["발췌 Category"])) for _, r in filtered_task_df.iterrows()}
    eq_nos = {k[0] for k in keys}
    filtered_cases = [c for c in repeat_cases if c.equipment_no in eq_nos]
    filtered_events = []
    for e in all_events:
        event_keys = {(e.equipment_no, cat) for cat in categorize_event(e)}
        if event_keys.intersection(keys):
            filtered_events.append(e)
    return filtered_cases, filtered_events


def _call_export_excel(
    task_df: pd.DataFrame,
    repeat_cases: list,
    all_events: list,
    output_path: Path,
    category_source_df: pd.DataFrame | None = None,
    extra_sheets: dict[str, pd.DataFrame] | None = None,
) -> None:
    try:
        export_excel(
            task_df,
            repeat_cases,
            all_events,
            output_path,
            category_source_df=category_source_df,
            extra_sheets=extra_sheets,
        )
    except TypeError:
        export_excel(task_df, repeat_cases, all_events, output_path)



def make_fixed_excel_bytes(
    task_df: pd.DataFrame,
    repeat_cases: list,
    all_events: list,
    filename_prefix: str,
    category_source_df: pd.DataFrame | None = None,
    extra_sheets: dict[str, pd.DataFrame] | None = None,
) -> tuple[bytes, str]:
    temp_dir = Path(tempfile.mkdtemp(prefix="repeat_task_export_"))
    output_path = temp_dir / f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    _call_export_excel(task_df, repeat_cases, all_events, output_path, category_source_df=category_source_df, extra_sheets=extra_sheets)
    _apply_readable_excel_format(output_path)
    return output_path.read_bytes(), output_path.name


def run_fixed_analysis(files: list) -> None:
    progress_bar = st.progress(0, text="분석 준비 중...")
    status_box = st.empty()

    usable_files, skipped_files = split_fixed_input_files(files)
    if not usable_files:
        raise ValueError("분석 가능한 원본 입력 파일이 없습니다. 기존 결과 파일은 제외하고 목록/Trouble List/원본 보고서만 업로드해 주세요.")

    temp_dir = save_uploaded_files(usable_files)
    status_box.info("고정장치 자료 분석 중...")

    def update_progress(current: int, total: int, filename: str):
        pct = int(current / total * 100) if total else 100
        progress_bar.progress(pct, text=f"분석 중... ({current}/{total}) {filename}")

    year_range = st.session_state.get("year_range", None)
    repeat_task_df, repeat_cases, all_events, equipment_names = run_pipeline_v6(temp_dir, progress_callback=update_progress, year_range=year_range)
    overview_df = build_equipment_summary_dataframe(all_events)
    full_excel_bytes, full_excel_name = make_fixed_excel_bytes(
        overview_df,
        repeat_cases,
        all_events,
        "반복정비_고정장치_전체결과_v6",
        category_source_df=overview_df,
        extra_sheets={"2회이상반복_과제후보": repeat_task_df},
    )

    progress_bar.progress(100, text="완료")
    status_box.success("고정장치 분석 완료")

    st.session_state["fixed_analysis_result"] = {
        "overview_df": overview_df,
        "repeat_task_df": repeat_task_df,
        "repeat_cases": repeat_cases,
        "all_events": all_events,
        "equipment_names": equipment_names,
        "full_excel_bytes": full_excel_bytes,
        "full_excel_name": full_excel_name,
        "uploaded_count": len(usable_files),
        "raw_uploaded_count": len(files),
        "skipped_files": skipped_files,
    }


# -------------------------------------------------
# piping helpers
# -------------------------------------------------
def _condense_df_for_display(df: pd.DataFrame, cols: list[str], text_cols: list[str], limit: int = 160) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=cols)
    out = df.copy()
    for c in text_cols:
        if c in out.columns:
            out[c] = out[c].astype(str).map(lambda x: _truncate_multiline(x, limit))
    use_cols = [c for c in cols if c in out.columns]
    return out[use_cols].copy()


def build_piping_candidate_df(repeat_df: pd.DataFrame) -> pd.DataFrame:
    if repeat_df is None or repeat_df.empty:
        return pd.DataFrame(columns=["NO", "Line Number", "발생구분", "발생횟수", "발생년도", "최근발생일", "대표조치", "검토메모"])
    out = repeat_df.copy().reset_index(drop=True)
    out["NO"] = range(1, len(out) + 1)
    return out


def filter_piping_repeat_df(repeat_df: pd.DataFrame, min_count: int, line_keyword: str, source_keyword: str) -> pd.DataFrame:
    if repeat_df is None or repeat_df.empty:
        return pd.DataFrame(columns=repeat_df.columns if repeat_df is not None else [])
    out = repeat_df.copy()
    out = out[pd.to_numeric(out["발생횟수"], errors="coerce").fillna(0) >= min_count].copy()
    if line_keyword:
        out = out[out["Line Number"].astype(str).str.contains(line_keyword, case=False, na=False)].copy()
    if source_keyword:
        out = out[out["출처"].astype(str).str.contains(source_keyword, case=False, na=False)].copy()
    out = out.reset_index(drop=True)
    out["NO"] = range(1, len(out) + 1)
    return out


def filter_piping_occurrences_by_lines(occurrences_df: pd.DataFrame, lines: set[str]) -> pd.DataFrame:
    if occurrences_df is None or occurrences_df.empty or not lines:
        return pd.DataFrame(columns=occurrences_df.columns if occurrences_df is not None else [])
    return occurrences_df[occurrences_df["line_no"].astype(str).isin(lines)].copy().reset_index(drop=True)


def filter_piping_excluded_by_lines(excluded_df: pd.DataFrame, lines: set[str], show_all: bool = False) -> pd.DataFrame:
    if excluded_df is None or excluded_df.empty:
        return pd.DataFrame(columns=excluded_df.columns if excluded_df is not None else [])
    if show_all or not lines:
        return excluded_df.copy().reset_index(drop=True)
    return excluded_df[excluded_df["line_no"].astype(str).isin(lines)].copy().reset_index(drop=True)


def make_piping_excel_bytes(summary_df: pd.DataFrame, repeat_df: pd.DataFrame, occurrences_df: pd.DataFrame, excluded_df: pd.DataFrame, filename_prefix: str) -> tuple[bytes, str]:
    temp_dir = Path(tempfile.mkdtemp(prefix="repeat_piping_export_"))
    output_path = temp_dir / f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    if export_dataframes is not None:
        export_dataframes(
            [
                ("summary", summary_df),
                ("반복배관후보", repeat_df),
                ("occurrences", occurrences_df),
                ("excluded_review", excluded_df),
            ],
            output_path,
        )
    else:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            summary_df.to_excel(writer, index=False, sheet_name="summary")
            repeat_df.to_excel(writer, index=False, sheet_name="반복배관후보")
            occurrences_df.to_excel(writer, index=False, sheet_name="occurrences")
            excluded_df.to_excel(writer, index=False, sheet_name="excluded_review")
    _apply_readable_excel_format(output_path)
    return output_path.read_bytes(), output_path.name


def run_piping_analysis(history_file, trouble_file) -> None:
    if not (build_piping_repeat_report_dataframe and build_piping_summary_sheet and load_piping_replacement_occurrences):
        raise RuntimeError("배관 분석 모듈이 패키지에 포함되지 않았습니다. 이번 패치본은 고정장치 Streamlit 개선 중심으로 재패키징되었습니다.")

    status_box = st.empty()
    progress_bar = st.progress(0, text="분석 준비 중...")
    status_box.info("배관 반복 교체 이력 분석 중...")

    temp_dir = Path(tempfile.mkdtemp(prefix="repeat_piping_input_"))
    hist_path = temp_dir / Path(history_file.name).name
    trb_path = temp_dir / Path(trouble_file.name).name
    hist_path.write_bytes(history_file.getbuffer())
    progress_bar.progress(20, text="배관 정비이력 저장 완료")
    trb_path.write_bytes(trouble_file.getbuffer())
    progress_bar.progress(40, text="Trouble List 저장 완료")

    occurrences_df, excluded_df = load_piping_replacement_occurrences(hist_path, trb_path)
    progress_bar.progress(75, text="배관 교체 occurrence 집계 완료")
    repeat_df = build_piping_repeat_report_dataframe(occurrences_df)
    summary_df = build_piping_summary_sheet(occurrences_df, excluded_df, repeat_df)
    full_excel_bytes, full_excel_name = make_piping_excel_bytes(
        summary_df=summary_df,
        repeat_df=repeat_df,
        occurrences_df=occurrences_df,
        excluded_df=excluded_df,
        filename_prefix="반복정비_배관_LineNumber_반복교체",
    )

    progress_bar.progress(100, text="완료")
    status_box.success("배관 분석 완료")

    st.session_state["piping_analysis_result"] = {
        "summary_df": summary_df,
        "repeat_df": repeat_df,
        "occurrences_df": occurrences_df,
        "excluded_df": excluded_df,
        "full_excel_bytes": full_excel_bytes,
        "full_excel_name": full_excel_name,
        "history_name": history_file.name,
        "trouble_name": trouble_file.name,
    }


# -------------------------------------------------
# sidebar / mode select
# -------------------------------------------------
with st.sidebar:
    st.header("실행 설정")
    mode = st.radio("분석 대상", options=["고정장치", "배관"], index=0)

    if mode == "고정장치":
        uploaded_files = st.file_uploader(
            "자료 업로드",
            type=["pdf", "xlsx", "xls", "xlsm", "xlsb", "txt", "csv"],
            accept_multiple_files=True,
            help="TA 보고서 PDF, Trouble List/정비 이력 Excel 등을 업로드하세요.",
            key="fixed_uploads",
        )
        st.subheader("검사 일자 범위")
        year_range = st.slider(
            "조회 기간 (검사일 기준)",
            min_value=1990,
            max_value=2030,
            value=(2000, 2026),
            step=1,
            key="year_range",
            help="선택한 기간 내의 검사 이력만 분석에 포함됩니다.",
        )
        pre_occurrence_option = st.radio("발생기준", options=["1회성 이상", "2회 이상 반복"], index=1, key="fixed_occurrence")
        pre_selected_categories = st.multiselect("카테고리", options=CATEGORY_ORDER, default=CATEGORY_ORDER, key="fixed_categories")
        run_btn = st.button("고정장치 분석 실행", type="primary", key="run_fixed")
    else:
        history_file = st.file_uploader(
            "배관 정비이력 파일",
            type=["xlsx", "xls", "xlsm", "xlsb"],
            accept_multiple_files=False,
            help="예: 목록_20260330092940.xlsx",
            key="piping_history",
        )
        trouble_file = st.file_uploader(
            "Trouble List 파일",
            type=["xlsx", "xls", "xlsm", "xlsb"],
            accept_multiple_files=False,
            help="예: 목록_20260330092956.xlsx",
            key="piping_trouble",
        )
        piping_occurrence_option = st.radio("반복기준", options=["1회성 교체 이상", "2회 이상 반복"], index=1, key="piping_occurrence")
        show_all_excluded = st.checkbox("제외 검토 시트 전체 보기", value=False, key="piping_show_all_excluded")
        run_btn = st.button("배관 분석 실행", type="primary", key="run_piping")


# -------------------------------------------------
# guide text
# -------------------------------------------------
if mode == "고정장치":
    st.markdown(
        """
### 사용 방법
1. 왼쪽에서 TA 보고서 / Trouble List / 정비이력 파일 업로드
2. 분석 전에 **발생기준 / 카테고리** 선택
3. **고정장치 분석 실행** 클릭
4. 결과는 **설비 + 카테고리** 기준으로 분리되어 표시됨

### 해석 기준
- **2회 이상 반복** = 같은 카테고리의 실제 조치가 서로 다른 연도에 2회 이상 수행된 경우
- recommendation만 있는 문장은 발생 건수에 포함하지 않음
- 내부 충진물(filler) 교체는 분류에서 제외
"""
    )
else:
    st.markdown(
        """
### 배관 분석 목적
1. **Line Number 기준**으로 실제 교체 이력을 수집
2. 동일 Line / 동일 일자 / 동일 성격 문장은 1건으로 중복 제거
3. **2회 이상 반복된 Line Number**를 반복 정비 후보로 도출

### 권장 입력 파일
1. 배관 정비이력 엑셀 1개
2. Trouble List 엑셀 1개

### 해석 기준
- **1회성 교체 이상** = 교체 occurrence 전체 검토
- **2회 이상 반복** = 동일 Line Number가 서로 다른 날짜에 2건 이상 발생
- Trouble History 요약행은 가능한 경우 과거 연도/날짜 단위로 분해하여 반영
"""
    )


# -------------------------------------------------
# run actions
# -------------------------------------------------
if run_btn and mode == "고정장치":
    if not uploaded_files:
        st.error("먼저 분석할 파일을 업로드해 주세요.")
        st.stop()
    if not pre_selected_categories:
        st.error("카테고리를 최소 1개 이상 선택해 주세요.")
        st.stop()
    try:
        st.session_state["pre_occurrence_option"] = pre_occurrence_option
        st.session_state["pre_selected_categories"] = pre_selected_categories
        run_fixed_analysis(uploaded_files)
    except Exception as exc:
        st.error(f"분석 중 오류 발생: {exc}")
        st.exception(exc)
        st.stop()

if run_btn and mode == "배관":
    if history_file is None or trouble_file is None:
        st.error("배관 정비이력 파일과 Trouble List 파일을 모두 업로드해 주세요.")
        st.stop()
    try:
        run_piping_analysis(history_file, trouble_file)
    except Exception as exc:
        st.error(f"분석 중 오류 발생: {exc}")
        st.exception(exc)
        st.stop()


# -------------------------------------------------
# fixed equipment result rendering
# -------------------------------------------------
fixed_result = st.session_state.get("fixed_analysis_result")
if mode == "고정장치" and fixed_result:
    overview_df = fixed_result["overview_df"]
    repeat_task_df = fixed_result["repeat_task_df"]
    repeat_cases = fixed_result["repeat_cases"]
    all_events = fixed_result["all_events"]

    one_time_count = int((pd.to_numeric(overview_df.get("발생년도수"), errors="coerce").fillna(0) == 1).sum()) if not overview_df.empty else 0
    repeat_count = int((pd.to_numeric(overview_df.get("발생년도수"), errors="coerce").fillna(0) >= 2).sum()) if not overview_df.empty else 0

    c1, c2, c3, c4, c5 = st.columns(5)
    raw_uploaded_count = fixed_result.get("raw_uploaded_count", fixed_result["uploaded_count"])
    c1.metric("입력 파일(사용/전체)", f"{fixed_result['uploaded_count']:,}/{raw_uploaded_count:,}")
    c2.metric("연도별 정비 이벤트 수", f"{len(all_events):,}")
    c3.metric("설비+카테고리 요약 수", f"{len(overview_df):,}")
    c4.metric("1회성", f"{one_time_count:,}")
    c5.metric("2회 이상 반복", f"{repeat_count:,}")

    skipped_files = fixed_result.get("skipped_files", [])
    if skipped_files:
        st.warning("이전 결과 파일은 입력에서 자동 제외했습니다: " + ", ".join(skipped_files))

    st.subheader("필터 결과")
    occurrence_option = st.radio(
        "발생기준 선택",
        options=["1회성 이상", "2회 이상 반복"],
        index=0 if st.session_state.get("pre_occurrence_option") == "1회성 이상" else 1,
        horizontal=True,
    )
    selected_categories = st.multiselect(
        "카테고리 재선택",
        options=CATEGORY_ORDER,
        default=st.session_state.get("pre_selected_categories", CATEGORY_ORDER),
    )
    min_year_count = 1 if occurrence_option == "1회성 이상" else 2

    base_task_df = repeat_task_df if (occurrence_option == "2회 이상 반복" and repeat_task_df is not None and not repeat_task_df.empty) else overview_df
    filtered_task_df = filter_task_df(base_task_df, selected_categories, min_year_count)
    filtered_cases, filtered_events = filter_related_objects(filtered_task_df, repeat_cases, all_events)
    category_extract_df = build_category_extract_dataframe(filtered_task_df)
    display_task_df = build_task_display_df(filtered_task_df)
    events_df = build_events_df(filtered_events)
    cases_df = build_cases_df(filtered_cases)

    filtered_excel_bytes, filtered_excel_name = make_fixed_excel_bytes(
        filtered_task_df,
        filtered_cases,
        filtered_events,
        "반복정비_고정장치_필터결과_v6",
        category_source_df=filtered_task_df,
    )

    st.caption(f"현재 조건: {occurrence_option} / {', '.join(selected_categories) if selected_categories else '카테고리 없음'}")
    col_dl1, col_dl2 = st.columns(2)
    with col_dl1:
        st.download_button(
            "현재 필터 결과 Excel 다운로드",
            data=filtered_excel_bytes,
            file_name=filtered_excel_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    with col_dl2:
        st.download_button(
            "전체 결과 Excel 다운로드",
            data=fixed_result["full_excel_bytes"],
            file_name=fixed_result["full_excel_name"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    tab1, tab2, tab3, tab4 = st.tabs(["과제후보_등록형식", "카테고리별_발췌", "반복정비_Case", "연도별_정비이벤트"])
    with tab1:
        st.dataframe(display_task_df, use_container_width=True, height=420)
    with tab2:
        st.dataframe(category_extract_df, use_container_width=True, height=360)
    with tab3:
        st.dataframe(cases_df, use_container_width=True, height=280)
    with tab4:
        st.dataframe(events_df, use_container_width=True, height=380)


# -------------------------------------------------
# piping result rendering
# -------------------------------------------------
piping_result = st.session_state.get("piping_analysis_result")
if mode == "배관" and piping_result:
    summary_df = piping_result["summary_df"]
    repeat_df = piping_result["repeat_df"]
    occurrences_df = piping_result["occurrences_df"]
    excluded_df = piping_result["excluded_df"]

    total_occ = len(occurrences_df) if occurrences_df is not None else 0
    total_repeat = len(repeat_df) if repeat_df is not None else 0
    excluded_total = len(excluded_df) if excluded_df is not None else 0
    source_mix_count = 0
    if occurrences_df is not None and not occurrences_df.empty and "source_count" in occurrences_df.columns:
        source_mix_count = int((pd.to_numeric(occurrences_df["source_count"], errors="coerce").fillna(0) >= 2).sum())

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("입력 파일", "2개")
    c2.metric("교체 occurrence", f"{total_occ:,}")
    c3.metric("반복 Line Number", f"{total_repeat:,}")
    c4.metric("양쪽 소스 중복확인", f"{source_mix_count:,}")
    c5.metric("제외 검토 행", f"{excluded_total:,}")

    st.caption(f"입력 파일: {piping_result['history_name']} / {piping_result['trouble_name']}")

    left, right = st.columns([2, 1])
    with left:
        min_count = 1 if st.session_state.get("piping_occurrence") == "1회성 교체 이상" else 2
        line_keyword = st.text_input("Line Number 필터", value="", placeholder="예: 12-PI-001 또는 01A", key="piping_line_filter")
        source_keyword = st.text_input("출처 필터", value="", placeholder="예: Trouble List 또는 배관 정비이력", key="piping_source_filter")
    with right:
        st.write("")
        st.write("")
        st.write(f"현재 기준: {'1회 이상 교체' if min_count == 1 else '2회 이상 반복'}")

    filtered_repeat_df = filter_piping_repeat_df(repeat_df, min_count=min_count, line_keyword=line_keyword, source_keyword=source_keyword)
    selected_lines = set(filtered_repeat_df["Line Number"].astype(str)) if not filtered_repeat_df.empty else set()

    filtered_occurrences_df = filter_piping_occurrences_by_lines(occurrences_df, selected_lines) if min_count >= 2 else occurrences_df.copy()
    filtered_excluded_df = filter_piping_excluded_by_lines(excluded_df, selected_lines, show_all=st.session_state.get("piping_show_all_excluded", False))

    filtered_summary_df = build_piping_summary_sheet(filtered_occurrences_df, filtered_excluded_df, filtered_repeat_df)
    filtered_excel_bytes, filtered_excel_name = make_piping_excel_bytes(
        summary_df=filtered_summary_df,
        repeat_df=filtered_repeat_df,
        occurrences_df=filtered_occurrences_df,
        excluded_df=filtered_excluded_df,
        filename_prefix="반복정비_배관_필터결과",
    )

    dl1, dl2 = st.columns(2)
    with dl1:
        st.download_button(
            "현재 필터 결과 Excel 다운로드",
            data=filtered_excel_bytes,
            file_name=filtered_excel_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    with dl2:
        st.download_button(
            "전체 결과 Excel 다운로드",
            data=piping_result["full_excel_bytes"],
            file_name=piping_result["full_excel_name"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    repeat_display_df = _condense_df_for_display(
        build_piping_candidate_df(filtered_repeat_df),
        cols=["NO", "Line Number", "발생구분", "발생횟수", "발생년도", "최근발생일", "출처", "대표조치", "상세이력", "검토메모"],
        text_cols=["대표조치", "상세이력", "검토메모"],
        limit=220,
    )
    occ_display_df = _condense_df_for_display(
        filtered_occurrences_df,
        cols=["line_no", "event_date", "event_year", "sources", "source_count", "titles", "details"],
        text_cols=["titles", "details"],
        limit=220,
    )
    ex_display_df = _condense_df_for_display(
        filtered_excluded_df,
        cols=["line_no", "event_date", "source_type", "title", "detail", "exclude_reason"],
        text_cols=["title", "detail", "exclude_reason"],
        limit=220,
    )

    tab1, tab2, tab3, tab4 = st.tabs(["반복배관후보", "교체_occurrences", "제외검토", "summary"])
    with tab1:
        st.dataframe(repeat_display_df, use_container_width=True, height=430)
    with tab2:
        st.dataframe(occ_display_df, use_container_width=True, height=420)
    with tab3:
        st.dataframe(ex_display_df, use_container_width=True, height=420)
    with tab4:
        st.dataframe(summary_df, use_container_width=True, height=240)
