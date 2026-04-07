from __future__ import annotations

from math import ceil
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .task_builder import build_category_extract_dataframe

TITLE_FILL = PatternFill(fill_type="solid", fgColor="17324D")
SUBTITLE_FILL = PatternFill(fill_type="solid", fgColor="DCEAF7")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="맑은 고딕", size=10)
TITLE_FONT = Font(color="FFFFFF", bold=True, name="맑은 고딕", size=15)
SUBTITLE_FONT = Font(color="17324D", bold=False, name="맑은 고딕", size=10)
BODY_FONT = Font(color="222222", name="맑은 고딕", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F2"),
    right=Side(style="thin", color="D9E2F2"),
    top=Side(style="thin", color="D9E2F2"),
    bottom=Side(style="thin", color="D9E2F2"),
)
ROW_FILL_EVEN = PatternFill(fill_type="solid", fgColor="F8FBFF")
ROW_FILL_ODD = PatternFill(fill_type="solid", fgColor="FFFFFF")
SHEET_TAB_COLORS = {
    "과제후보_등록형식": "17324D",
    "카테고리별_발췌": "2F75B5",
    "반복정비_Case": "5B9BD5",
    "연도별_정비이벤트": "70AD47",
    "요약_대시보드": "7F60A8",
}
CATEGORY_FILL_MAP = {
    "단순 보수": "EAF2F8",
    "도장": "FDEBD0",
    "육성용접": "FADBD8",
    "단순 내부 구성품 교체": "E8F8F5",
    "Nozzle 교체": "EBDEF0",
    "Assembly 교체": "D6EAF8",
}
CENTER_HEADERS = {
    "NO", "Equipment No", "발생구분", "발생년도수", "발생년도", "발췌 Category",
    "검토필요여부", "year", "event_year", "event_date", "source_count", "confidence",
    "발생횟수", "최근발생일", "source_type",
}
WIDE_TEXT_HEADERS = {
    "설비명", "TA 조치사항", "추후 권고사항", "제목", "상세 내용", "source_files",
    "finding_location", "finding_damage", "finding_measurement", "action_type", "action_detail",
    "recommendation", "evidence_summary", "repeat_reason", "action_cluster", "location_cluster",
    "damage_cluster", "sources", "titles", "details", "출처", "대표조치", "상세이력",
    "검토메모", "title", "detail", "exclude_reason",
}
FIXED_WIDTHS = {
    "NO": 7,
    "Equipment No": 15,
    "설비명": 28,
    "발생구분": 13,
    "발생년도수": 11,
    "발생년도": 18,
    "발췌 Category": 18,
    "TA 조치사항": 60,
    "추후 권고사항": 60,
    "제목": 38,
    "상세 내용": 92,
    "검토필요여부": 12,
    "equipment_no": 15,
    "equipment_name": 28,
    "year": 10,
    "source_files": 32,
    "finding_location": 22,
    "finding_damage": 24,
    "finding_measurement": 20,
    "action_type": 22,
    "action_detail": 64,
    "recommendation": 64,
    "evidence_summary": 82,
    "action_cluster": 20,
    "location_cluster": 20,
    "damage_cluster": 20,
    "repeat_reason": 44,
    "confidence": 10,
}


def _safe_df(df) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame()
    return df.copy()


def _stringify(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return str(value)


def _visible_len(value) -> int:
    text = _stringify(value)
    if not text:
        return 0
    return max((len(line) for line in text.splitlines()), default=0)


def _estimate_width(ws, col_idx: int, header: str, header_row: int) -> int:
    if header in FIXED_WIDTHS:
        return FIXED_WIDTHS[header]
    max_len = len(header)
    max_row = min(ws.max_row, header_row + 200)
    for row_idx in range(header_row + 1, max_row + 1):
        max_len = max(max_len, _visible_len(ws.cell(row=row_idx, column=col_idx).value))
    if header in WIDE_TEXT_HEADERS:
        return max(24, min(62, int(max_len * 1.05) + 2))
    return max(10, min(26, int(max_len * 1.08) + 2))


def _estimated_wrapped_lines(text: str, width: int) -> int:
    if not text:
        return 1
    usable = max(10, width - 1)
    total = 0
    for raw_line in text.splitlines() or [text]:
        line = raw_line.strip() or " "
        total += max(1, ceil(len(line) / usable))
    return total


def _sheet_subtitle(sheet_name: str, df: pd.DataFrame) -> str:
    return f"행 수 {len(df):,} / 다운로드 가독성 최적화 버전"


def _style_title_block(ws, last_col: int, title: str, subtitle: str) -> int:
    end_col = get_column_letter(max(1, last_col))
    ws.merge_cells(f"A1:{end_col}1")
    ws.merge_cells(f"A2:{end_col}2")
    ws["A1"] = title
    ws["A2"] = subtitle
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A2"].fill = SUBTITLE_FILL
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 20
    return 4


def _apply_category_fill(cell) -> None:
    category = _stringify(cell.value)
    color = CATEGORY_FILL_MAP.get(category)
    if color:
        cell.fill = PatternFill(fill_type="solid", fgColor=color)
        cell.font = Font(color="17324D", bold=True, name="맑은 고딕", size=10)


def _apply_sheet_formatting(ws, sheet_name: str, df: pd.DataFrame) -> None:
    if ws.max_column == 0:
        return

    header_row = _style_title_block(ws, ws.max_column, sheet_name, _sheet_subtitle(sheet_name, df))
    ws.freeze_panes = f"A{header_row + 1}"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90

    headers = []
    category_col = None
    confidence_col = None
    review_col = None

    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        header = _stringify(cell.value)
        headers.append(header)
        if header == "발췌 Category":
            category_col = col_idx
        if header == "confidence":
            confidence_col = col_idx
        if header == "검토필요여부":
            review_col = col_idx
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = _estimate_width(ws, col_idx, header, header_row)
    ws.row_dimensions[header_row].height = 24

    if ws.max_row >= header_row:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
        if ws.max_row > header_row:
            table = Table(displayName=f"T_{abs(hash(sheet_name))}"[:31], ref=f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}")
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
            ws.add_table(table)

    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_fill = ROW_FILL_EVEN if (row_idx - header_row) % 2 == 0 else ROW_FILL_ODD
        row_line_estimate = 1
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            text = _stringify(cell.value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if not cell.fill or cell.fill.fill_type is None:
                cell.fill = row_fill
            width = int(ws.column_dimensions[get_column_letter(col_idx)].width or 12)
            if header in CENTER_HEADERS:
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if header in WIDE_TEXT_HEADERS or len(text) > 24 or "\n" in text:
                row_line_estimate = max(row_line_estimate, _estimated_wrapped_lines(text, width))
        if category_col:
            _apply_category_fill(ws.cell(row=row_idx, column=category_col))
        if review_col:
            rcell = ws.cell(row=row_idx, column=review_col)
            value = _stringify(rcell.value)
            if value in {"Y", "예", "필요", "검토필요"}:
                rcell.fill = PatternFill(fill_type="solid", fgColor="FDE9D9")
                rcell.font = Font(color="9C0006", bold=True, name="맑은 고딕", size=10)
            elif value:
                rcell.fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
        if confidence_col:
            ccell = ws.cell(row=row_idx, column=confidence_col)
            try:
                score = float(ccell.value)
            except Exception:
                score = None
            if score is not None:
                if score >= 0.85:
                    ccell.fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
                elif score >= 0.6:
                    ccell.fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
                else:
                    ccell.fill = PatternFill(fill_type="solid", fgColor="FDE9D9")
        long_text_boost = 0
        if row_line_estimate >= 5:
            long_text_boost = 10
        if row_line_estimate >= 8:
            long_text_boost = 18
        ws.row_dimensions[row_idx].height = min(220, max(22, 18 * row_line_estimate + 6 + long_text_boost))

    ws.sheet_properties.tabColor = SHEET_TAB_COLORS.get(sheet_name, "5B9BD5")
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = f"{header_row}:{header_row}"


def _create_dashboard(workbook, frames: dict[str, pd.DataFrame]) -> None:
    ws = workbook.create_sheet("요약_대시보드", 0)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.sheet_properties.tabColor = SHEET_TAB_COLORS["요약_대시보드"]

    ws.merge_cells("A1:H1")
    ws["A1"] = "반복정비 결과 요약 대시보드"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = Font(color="FFFFFF", bold=True, name="맑은 고딕", size=16)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:H2")
    ws["A2"] = "다운로드 즉시 보기 편하도록 핵심 지표, 카테고리 분포, 시트 안내를 한 장에 요약했습니다."
    ws["A2"].fill = SUBTITLE_FILL
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22

    task_df = frames.get("과제후보_등록형식", pd.DataFrame())
    category_df = frames.get("카테고리별_발췌", pd.DataFrame())
    case_df = frames.get("반복정비_Case", pd.DataFrame())
    event_df = frames.get("연도별_정비이벤트", pd.DataFrame())

    metrics = [
        ("과제 후보", len(task_df), "핵심 등록형 과제 수"),
        ("카테고리 발췌", len(category_df), "검토용 발췌 총행"),
        ("반복정비 Case", len(case_df), "설비 반복 이슈 묶음"),
        ("연도별 이벤트", len(event_df), "연도별 정비 이벤트 수"),
    ]

    start_col = 1
    for idx, (label, value, note) in enumerate(metrics):
        col = start_col + idx * 2
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
        ws.cell(4, col).value = label
        ws.cell(4, col).fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        ws.cell(4, col).font = Font(color="FFFFFF", bold=True, name="맑은 고딕", size=10)
        ws.cell(4, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=5, start_column=col, end_row=6, end_column=col + 1)
        ws.cell(5, col).value = f"{value:,}"
        ws.cell(5, col).fill = PatternFill(fill_type="solid", fgColor="F8FBFF")
        ws.cell(5, col).font = Font(color="17324D", bold=True, name="맑은 고딕", size=18)
        ws.cell(5, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col + 1)
        ws.cell(7, col).value = note
        ws.cell(7, col).fill = PatternFill(fill_type="solid", fgColor="EDF4FB")
        ws.cell(7, col).font = Font(color="44546A", name="맑은 고딕", size=9)
        ws.cell(7, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in range(4, 8):
        ws.row_dimensions[r].height = 22

    ws["A10"] = "시트 바로가기"
    ws["A10"].font = Font(color="17324D", bold=True, name="맑은 고딕", size=12)
    ws["A10"].fill = PatternFill(fill_type="solid", fgColor="DCEAF7")
    ws["A11"] = "시트명"
    ws["B11"] = "설명"
    for c in ["A11", "B11"]:
        ws[c].fill = HEADER_FILL
        ws[c].font = HEADER_FONT
        ws[c].alignment = Alignment(horizontal="center", vertical="center")
        ws[c].border = THIN_BORDER

    guides = [
        ("과제후보_등록형식", "최종 등록용 과제 후보 목록"),
        ("카테고리별_발췌", "카테고리 검토용 상세 발췌"),
        ("반복정비_Case", "반복 Case 묶음 및 반복 사유"),
        ("연도별_정비이벤트", "연도별 finding/action/recommendation 이벤트"),
    ]
    row = 12
    for sheet_name, desc in guides:
        ws.cell(row, 1).value = sheet_name
        ws.cell(row, 1).hyperlink = f"#{sheet_name}!A1"
        ws.cell(row, 1).style = "Hyperlink"
        ws.cell(row, 2).value = desc
        for col in (1, 2):
            ws.cell(row, col).border = THIN_BORDER
            ws.cell(row, col).fill = ROW_FILL_EVEN if row % 2 == 0 else ROW_FILL_ODD
            ws.cell(row, col).alignment = Alignment(vertical="center", wrap_text=True)
            ws.cell(row, col).font = BODY_FONT
        row += 1

    ws["D10"] = "카테고리 분포"
    ws["D10"].font = Font(color="17324D", bold=True, name="맑은 고딕", size=12)
    ws["D10"].fill = PatternFill(fill_type="solid", fgColor="DCEAF7")
    cat_summary = pd.DataFrame()
    if not task_df.empty and "발췌 Category" in task_df.columns:
        cat_summary = (
            task_df.groupby("발췌 Category", dropna=False)
            .agg(건수=("Equipment No", "count"), 설비수=("Equipment No", "nunique"))
            .reset_index()
            .sort_values(["건수", "설비수"], ascending=[False, False])
        )
    ws["D11"] = "발췌 Category"
    ws["E11"] = "건수"
    ws["F11"] = "설비수"
    for c in ["D11", "E11", "F11"]:
        ws[c].fill = HEADER_FILL
        ws[c].font = HEADER_FONT
        ws[c].alignment = Alignment(horizontal="center", vertical="center")
        ws[c].border = THIN_BORDER
    row = 12
    for _, rec in cat_summary.iterrows():
        ws.cell(row, 4).value = rec["발췌 Category"]
        ws.cell(row, 5).value = int(rec["건수"])
        ws.cell(row, 6).value = int(rec["설비수"])
        for col in range(4, 7):
            cell = ws.cell(row, col)
            cell.border = THIN_BORDER
            cell.fill = ROW_FILL_EVEN if row % 2 == 0 else ROW_FILL_ODD
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center" if col >= 5 else "left", vertical="center")
        _apply_category_fill(ws.cell(row, 4))
        row += 1

    ws["H10"] = "보기 팁"
    ws["H10"].font = Font(color="17324D", bold=True, name="맑은 고딕", size=12)
    ws["H10"].fill = PatternFill(fill_type="solid", fgColor="DCEAF7")
    tips = [
        "상단 요약 시트에서 전체 규모를 먼저 확인",
        "'과제후보_등록형식'은 바로 등록 검토용",
        "'카테고리별_발췌'는 문장 단위 비교 검토용",
        "필터와 고정틀이 적용되어 바로 탐색 가능",
    ]
    for idx, tip in enumerate(tips, start=11):
        ws.cell(idx, 8).value = f"• {tip}"
        ws.cell(idx, 8).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(idx, 8).font = BODY_FONT

    for col, width in {"A": 20, "B": 34, "D": 22, "E": 10, "F": 10, "H": 34}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A4"


def export_dataframes(sheet_frames: Iterable[tuple[str, pd.DataFrame]], output_path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    normalized_frames = [(str(name)[:31] or "Sheet1", _safe_df(df)) for name, df in sheet_frames]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in normalized_frames:
            df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=3)
        workbook = writer.book
        frames_map = {name: df for name, df in normalized_frames}
        _create_dashboard(workbook, frames_map)
        for sheet_name, df in normalized_frames:
            _apply_sheet_formatting(workbook[sheet_name], sheet_name, df)
    return output_path


def export_excel(task_df, repeat_cases, all_events, output_path, category_source_df=None, extra_sheets: dict[str, pd.DataFrame] | None = None):
    task_df = _safe_df(task_df)
    category_source_df = task_df if category_source_df is None else _safe_df(category_source_df)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    case_rows = []
    for c in repeat_cases or []:
        case_rows.append({
            "equipment_no": c.equipment_no,
            "설비명": c.equipment_name,
            "발생년도수": len(sorted(set(c.years))),
            "발생년도": ", ".join(map(str, sorted(set(c.years)))),
            "action_cluster": c.action_cluster,
            "location_cluster": c.location_cluster,
            "damage_cluster": c.damage_cluster,
            "repeat_reason": c.repeat_reason,
            "confidence": c.confidence,
        })
    cases_df = pd.DataFrame(case_rows)

    event_rows = []
    for e in all_events or []:
        event_rows.append({
            "equipment_no": e.equipment_no,
            "equipment_name": e.equipment_name,
            "year": e.report_year,
            "source_files": ", ".join(e.source_files or []),
            "finding_location": e.finding_location,
            "finding_damage": e.finding_damage,
            "finding_measurement": e.finding_measurement,
            "action_type": e.action_type,
            "action_detail": e.action_detail,
            "recommendation": e.recommendation,
            "evidence_summary": e.evidence_summary,
        })
    events_df = pd.DataFrame(event_rows)
    category_df = build_category_extract_dataframe(category_source_df)

    sheet_frames: list[tuple[str, pd.DataFrame]] = [
        ("과제후보_등록형식", task_df),
        ("카테고리별_발췌", category_df),
        ("반복정비_Case", cases_df),
        ("연도별_정비이벤트", events_df),
    ]
    for sheet_name, df in (extra_sheets or {}).items():
        sheet_frames.append((sheet_name, _safe_df(df)))

    return export_dataframes(sheet_frames, output_path)
